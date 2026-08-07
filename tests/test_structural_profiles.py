import json
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from ingest.mapper import build_column_map, map_row
from ingest.mappers.versioned import load_mapping
from ingest.probe import probe_file
from ingest.profiles import ProfileRegistry, load_profile_registry, make_candidate_profile
from ingest.structural import (
    FINGERPRINT_VERSION,
    ReaderLimits,
    StructuralError,
    classify_structure,
    fingerprint_file,
    group_by_fingerprint,
    inspect_file,
)


def _write_machine_csv(path: Path, rows: list[str], delimiter: str = ",") -> None:
    path.write_text(
        delimiter.join(("Timestamp", "Cycle_Time", "Machine")) + "\n" + "\n".join(rows) + "\n",
        encoding="utf-8",
    )


def test_fingerprint_uses_structure_not_values_or_file_name(tmp_path: Path):
    first = tmp_path / "customer-a.csv"
    second = tmp_path / "customer-b.csv"
    _write_machine_csv(first, ["2026-01-01T00:00:00Z,1.0,M1"])
    _write_machine_csv(second, ["2030-12-31T23:59:59Z,99.0,SECRET"])

    first_fp = fingerprint_file(first)
    second_fp = fingerprint_file(second)

    assert first_fp.digest == second_fp.digest
    serialized = json.dumps(first_fp.payload)
    assert "customer" not in serialized
    assert "SECRET" not in serialized
    assert "2026" not in serialized


def test_fingerprint_changes_for_column_order_and_delimiter(tmp_path: Path):
    comma = tmp_path / "comma.csv"
    semicolon = tmp_path / "semicolon.csv"
    reordered = tmp_path / "reordered.csv"
    _write_machine_csv(comma, ["2026-01-01T00:00:00Z,1.0,M1"])
    _write_machine_csv(semicolon, ["2026-01-01T00:00:00Z;1.0;M1"], delimiter=";")
    reordered.write_text("Cycle_Time,Timestamp,Machine\n1.0,2026-01-01T00:00:00Z,M1\n", encoding="utf-8")

    assert fingerprint_file(comma).digest != fingerprint_file(semicolon).digest
    assert fingerprint_file(comma).digest != fingerprint_file(reordered).digest


def test_utf16_transposed_table_is_detected(tmp_path: Path):
    source = tmp_path / "transposed.txt"
    source.write_text(
        "Variable\t1\t2\n"
        "Date\t2026-01-01\t2026-01-01\n"
        "Heure\t08:00:00\t08:00:01\n"
        "CycleTime\t1.0\t1.1\n",
        encoding="utf-16",
    )

    inspection = inspect_file(source)

    assert inspection.encoding == "utf-16"
    assert inspection.orientation == "transposed"
    assert inspection.columns == ("Date", "Heure", "CycleTime")
    assert inspection.rows_observed == 2
    assert probe_file(source)["validation"]["valid"] is True


def test_utf16_without_bom_is_treated_as_text(tmp_path: Path):
    source = tmp_path / "machine.txt"
    source.write_bytes(
        "Timestamp;Cycle_Time\n2026-01-01T00:00:00Z;1.0\n".encode("utf-16-le")
    )

    inspection = inspect_file(source)

    assert inspection.family == "delimited_text"
    assert inspection.encoding == "utf-16-le"
    assert inspection.columns == ("Timestamp", "Cycle_Time")


def test_normal_wide_csv_is_not_mistaken_for_transposed(tmp_path: Path):
    source = tmp_path / "wide.csv"
    headers = ["Timestamp"] + [f"Metric_{index}" for index in range(39)]
    row = ["2026-01-01T00:00:00Z"] + [str(index) for index in range(39)]
    source.write_text(",".join(headers) + "\n" + ",".join(row) + "\n" + ",".join(row) + "\n", encoding="utf-8")

    assert inspect_file(source).orientation == "rows"


def test_csv_preserves_empty_header_positions_and_multiline_fields(tmp_path: Path):
    source = tmp_path / "quoted.csv"
    source.write_text(
        'Timestamp,,Cycle_Time,Comment\n'
        '2026-01-01T00:00:00Z,unused,1.5,"line one\nline two"\n',
        encoding="utf-8",
    )

    inspection = inspect_file(source)

    assert inspection.columns == ("Timestamp", "__unnamed_1", "Cycle_Time", "Comment")
    assert inspection.rows_observed == 1
    assert inspection.sample_rows[0][2] == "1.5"
    assert inspection.sample_rows[0][3] == "line one\nline two"


def test_xlsx_selects_tabular_sheet_and_never_evaluates_formula(tmp_path: Path):
    source = tmp_path / "multi-sheet.xlsx"
    workbook = Workbook()
    cover = workbook.active
    cover.title = "Cover"
    cover["A1"] = "Summary"
    data = workbook.create_sheet("Data")
    data.append(["Timestamp", "Cycle_Time", "Machine"])
    data.append(["2026-01-01T00:00:00Z", "=1+1", "M1"])
    workbook.save(source)

    inspection = inspect_file(source)

    assert inspection.sheet_name == "Data"
    assert inspection.sheet_count == 2
    assert inspection.columns == ("Timestamp", "Cycle_Time", "Machine")
    assert "=1+1" not in {str(value) for row in inspection.sample_rows for value in row}
    assert inspection.sample_rows[0][1] is None


def test_xlsx_declared_wide_dimension_is_rejected_before_row_loading(tmp_path: Path):
    source = tmp_path / "wide.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Timestamp"
    sheet["XFD1"] = "TooWide"
    workbook.save(source)

    with pytest.raises(StructuralError) as error:
        inspect_file(source, limits=ReaderLimits(max_columns=100))

    assert error.value.code == "INGEST_TOO_MANY_COLUMNS"


def test_xlsx_explicit_missing_sheet_has_stable_error(tmp_path: Path):
    source = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active.append(["Timestamp", "Cycle_Time"])
    workbook.save(source)

    with pytest.raises(StructuralError) as error:
        inspect_file(source, sheet_name="Missing")

    assert error.value.code == "INGEST_SHEET_NOT_FOUND"
    assert str(source) not in str(error.value)


def test_xlsx_rejects_unsafe_member_compression_ratio(tmp_path: Path):
    source = tmp_path / "unsafe.xlsx"
    import zipfile

    with zipfile.ZipFile(source, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("xl/worksheets/sheet1.xml", b"0" * 100_000)

    with pytest.raises(StructuralError) as error:
        inspect_file(source, limits=ReaderLimits(max_compression_ratio=10))

    assert error.value.code == "INGEST_ARCHIVE_RATIO"


def test_legacy_xls_reader_is_bounded_and_explicit(tmp_path: Path, monkeypatch):
    source = tmp_path / "legacy.xls"
    source.write_bytes(b"synthetic-ole-placeholder")

    class FakeBook:
        sheet_names = ["Data"]

    monkeypatch.setattr(pd, "ExcelFile", lambda *args, **kwargs: FakeBook())
    monkeypatch.setattr(
        pd,
        "read_excel",
        lambda *args, **kwargs: pd.DataFrame([
            ["Timestamp", "Cycle_Time", "Machine"],
            ["2026-01-01T00:00:00Z", 1.0, "M1"],
        ]),
    )

    inspection = inspect_file(source)

    assert inspection.source_format == "xls"
    assert inspection.family == "spreadsheet"
    assert inspection.columns == ("Timestamp", "Cycle_Time", "Machine")


def test_binary_dat_is_classified_without_parsing_values(tmp_path: Path):
    source = tmp_path / "machine.dat"
    source.write_bytes(b"\x00\x01\x02\x03" * 100)

    inspection = inspect_file(source)
    classification = classify_structure(inspection)
    fingerprint = fingerprint_file(source)

    assert inspection.family == "binary"
    assert inspection.sample_rows == ()
    assert classification.label == "binary_unsupported"
    assert fingerprint.version == FINGERPRINT_VERSION
    assert fingerprint.payload["classification"] == "unsupported"


def test_non_dat_binary_is_rejected(tmp_path: Path):
    source = tmp_path / "machine.txt"
    source.write_bytes(b"\x00\x01\x02\x03" * 100)

    with pytest.raises(StructuralError) as error:
        inspect_file(source)

    assert error.value.code == "INGEST_BINARY_TEXT"


def test_reader_limits_file_columns_and_cells(tmp_path: Path):
    source = tmp_path / "large.csv"
    source.write_text("A,B,C\n1,2,3\n", encoding="utf-8")
    with pytest.raises(StructuralError) as error:
        inspect_file(source, limits=ReaderLimits(max_file_bytes=2))
    assert error.value.code == "INGEST_FILE_TOO_LARGE"

    with pytest.raises(StructuralError) as error:
        inspect_file(source, limits=ReaderLimits(max_columns=2))
    assert error.value.code == "INGEST_TOO_MANY_COLUMNS"

    cell = tmp_path / "cell.csv"
    cell.write_text("Timestamp,Cycle_Time\n2026-01-01T00:00:00Z,123456\n", encoding="utf-8")
    with pytest.raises(StructuralError) as error:
        inspect_file(cell, limits=ReaderLimits(max_cell_chars=5))
    assert error.value.code == "INGEST_CELL_TOO_LARGE"


def test_machine_classification_requires_time_and_process_metric(tmp_path: Path):
    machine = tmp_path / "machine.csv"
    erp = tmp_path / "erp.csv"
    _write_machine_csv(machine, ["2026-01-01T00:00:00Z,1.0,M1"])
    erp.write_text("Order,Machine,Quantity\nOF1,M1,10\n", encoding="utf-8")

    assert classify_structure(inspect_file(machine)).label == "machine_export"
    assert classify_structure(inspect_file(erp)).label == "tabular_unknown"


def test_group_by_fingerprint_clusters_same_structure(tmp_path: Path):
    first = tmp_path / "one.csv"
    second = tmp_path / "two.csv"
    third = tmp_path / "three.csv"
    _write_machine_csv(first, ["2026-01-01T00:00:00Z,1.0,M1"])
    _write_machine_csv(second, ["2026-01-02T00:00:00Z,2.0,M2"])
    third.write_text("Order,Quantity\nOF1,10\n", encoding="utf-8")

    groups = group_by_fingerprint([first, second, third])

    assert sorted(len(paths) for paths in groups.values()) == [1, 2]


def test_default_registry_matches_repository_candidate_profiles():
    arburg = fingerprint_file("data/samples/arburg_1003_cycles.txt")
    trs = fingerprint_file("data/samples/erp_trs_fevrier.xlsx")
    registry = load_profile_registry()

    arburg_match = registry.match(arburg)
    trs_match = registry.match(trs)

    assert arburg_match.status == "candidate"
    assert arburg_match.profile_id == "arburg-selogica-gestica-sample-v1"
    assert trs_match.status == "candidate"
    assert trs_match.profile_id == "erp-trs-sample-v1"
    assert trs_match.record_kind == "production_order"

    report = probe_file("data/samples/erp_trs_fevrier.xlsx")
    assert report["mapping"]["parser_version"] == "erp-trs-v1"
    assert len(report["mapping"]["recognized_columns"]) == 17
    assert report["validation"]["missing_required_fields"] == []
    assert report["validation"]["valid"] is True


def test_candidate_profile_collisions_fail_closed(tmp_path: Path):
    source = tmp_path / "machine.csv"
    _write_machine_csv(source, ["2026-01-01T00:00:00Z,1.0,M1"])
    fingerprint = fingerprint_file(source)
    first = make_candidate_profile(fingerprint, profile_id="candidate-one", parser_version="generic-v1")
    second = make_candidate_profile(fingerprint, profile_id="candidate-two", parser_version="generic-v1")

    with pytest.raises(StructuralError) as error:
        ProfileRegistry((first, second)).match(fingerprint)

    assert error.value.code == "INGEST_PROFILE_DUPLICATE"


def test_invalid_approved_profile_is_rejected(tmp_path: Path):
    profile_root = tmp_path / "profiles"
    profile_root.mkdir()
    (profile_root / "invalid.json").write_text(json.dumps({
        "schema": "iddrv.structural-profile",
        "schema_version": 1,
        "profile_id": "invalid-approved",
        "profile_version": "1.0.0",
        "status": "approved",
        "match_enabled": False,
        "fingerprint": None,
        "mapping": {"parser_version": "generic-v1"},
        "qualification": {"field_validated": False},
        "approval": {"approved_at": None, "approved_by": None},
    }), encoding="utf-8")

    with pytest.raises(StructuralError) as error:
        load_profile_registry(profile_root)

    assert error.value.code == "INGEST_PROFILE_INVALID"


def test_disabled_profiles_do_not_match_binary_dat(tmp_path: Path):
    source = tmp_path / "fanuc.dat"
    source.write_bytes(b"\x00\x01\x02" * 100)

    match = load_profile_registry().match(fingerprint_file(source))

    assert match.status == "none"


def test_approved_profile_wins_over_candidate(tmp_path: Path):
    source = tmp_path / "machine.csv"
    _write_machine_csv(source, ["2026-01-01T00:00:00Z,1.0,M1"])
    fingerprint = fingerprint_file(source)
    candidate = make_candidate_profile(fingerprint, profile_id="candidate-v1", parser_version="generic-v1")
    approved = dict(candidate)
    approved.update({
        "profile_id": "approved-v1",
        "status": "approved",
        "qualification": {"field_validated": True},
        "approval": {"approved_at": "2026-07-14T00:00:00Z", "approved_by": "test-reviewer"},
    })

    match = ProfileRegistry((candidate, approved)).match(fingerprint)

    assert match.status == "approved"
    assert match.profile_id == "approved-v1"


def test_probe_converts_units_declared_on_unit_row(tmp_path: Path):
    source = tmp_path / "units.txt"
    source.write_text(
        "Timestamp;Cycle_Time;Inj_Pres\n"
        "-;ms;psi\n"
        "2026-01-01T00:00:00Z;2500;2617\n",
        encoding="utf-8",
    )

    report = probe_file(source)
    recognized = {item["canonical"]: item for item in report["mapping"]["recognized_columns"]}

    assert recognized["cycle_time_s"]["source_unit"] == "ms"
    assert recognized["cycle_time_s"]["conversion_factor"] == pytest.approx(0.001)
    assert recognized["peak_pressure_bar"]["source_unit"] == "psi"
    assert recognized["peak_pressure_bar"]["conversion_factor"] == pytest.approx(0.0689475729)
    assert report["validation"]["unit_issues"] == []
    assert report["validation"]["valid"] is True


def test_explicit_unit_row_overrides_conflicting_header_suffix(tmp_path: Path):
    source = tmp_path / "units.csv"
    source.write_text(
        "Timestamp,Cycle_Time_ms\n"
        "-,s\n"
        "2026-01-01T00:00:00Z,2.5\n",
        encoding="utf-8",
    )

    report = probe_file(source)
    cycle = next(item for item in report["mapping"]["recognized_columns"] if item["canonical"] == "cycle_time_s")

    assert cycle["source_unit"] == "s"
    assert cycle["conversion_factor"] == 1.0
    assert cycle["conversion_applied"] is False


def test_sparse_unit_row_and_celsius_alias_are_supported(tmp_path: Path):
    sparse = tmp_path / "sparse.csv"
    sparse.write_text(
        "Timestamp,Vendor,Cycle_Time\n"
        ",,ms\n"
        "2026-01-01T00:00:00Z,x,2500\n",
        encoding="utf-8",
    )
    celsius = tmp_path / "temperature.csv"
    celsius.write_text(
        "Timestamp,oil_temperature_c\n"
        "-,°C\n"
        "2026-01-01T00:00:00Z,42\n",
        encoding="utf-8",
    )

    sparse_report = probe_file(sparse)
    celsius_report = probe_file(celsius)
    cycle = next(item for item in sparse_report["mapping"]["recognized_columns"] if item["canonical"] == "cycle_time_s")
    temperature = next(item for item in celsius_report["mapping"]["recognized_columns"] if item["canonical"] == "oil_temperature_c")

    assert sparse_report["profile"]["unit_row"] == 1
    assert cycle["conversion_factor"] == pytest.approx(0.001)
    assert temperature["unit_compatible"] is True
    assert celsius_report["validation"]["unit_issues"] == []


def test_probe_is_bounded_and_reports_structural_profile(tmp_path: Path):
    source = tmp_path / "machine.csv"
    _write_machine_csv(
        source,
        [f"2026-01-01T00:00:{index:02d}Z,1.{index},M1" for index in range(5)],
    )

    report = probe_file(source, limits=ReaderLimits(max_sample_rows=2))

    assert report["structure"]["fingerprint_version"] == FINGERPRINT_VERSION
    assert report["structure"]["classification"]["label"] == "machine_export"
    assert report["structure"]["bounded"] == {
        "rows_observed": 2,
        "max_rows": 2,
        "truncated": True,
    }
    assert report["validation"]["scope"] == "sample"
    assert report["validation"]["complete"] is False


def test_erp_profile_rejects_invalid_dates_numbers_and_ratios(tmp_path: Path):
    source = tmp_path / "erp.xlsx"
    headers = [
        "Réf OF", "Réf. Machine", "Lib. Machine", "Num Equipe", "Début Equipe",
        "Fin Equipe", "Réf. produit", "Lib. Produit", "Réf. outil", "Réf. Matière",
        "Tps Disponible (h)", "Tps Fct Brut (h)", "T.R.S.", "Cycle Moyen",
        "Nb Cycles", "Qté Pieces Bonnes", "Total Rebuts",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Données_Audit"
    sheet.append(headers)
    sheet.append([
        "OF1", "M1", "Machine", 1, "not-a-date", "2026-01-01 10:00", "P1", "Product",
        "T1", "MAT1", -1, "invalid-number", 1.5, 2.5, 10, 9, 1,
    ])
    workbook.save(source)

    report = probe_file(source)
    codes = {(item["field"], item["code"]) for item in report["validation"]["invalid_values"]}

    assert ("started_at", "invalid_timestamp") in codes
    assert ("erp_available_time_h", "negative_value") in codes
    assert ("erp_running_time_h", "invalid_numeric") in codes
    assert ("erp_trs", "ratio_out_of_range") in codes
    assert report["validation"]["valid"] is False


def test_mapping_preserves_numeric_identifiers_and_unknown_values():
    mapping = {
        "Order": {"canonical": "production_order_id", "scale": 1.0},
        "Machine": {"canonical": "machine_erp_ref", "scale": 1.0},
        "VendorCode": {"canonical": None},
    }

    row = map_row({"Order": "000123", "Machine": "001", "VendorCode": "0007"}, mapping)

    assert row["production_order_id"] == "000123"
    assert row["machine_erp_ref"] == "001"
    assert row["raw_data"]["VendorCode"] == "0007"


def test_mapping_selectors_cannot_escape_registry():
    with pytest.raises(ValueError, match="unsafe mapping selector"):
        load_mapping(machine_erp_ref="../../confidential", parser_version="generic-v1")
    with pytest.raises(ValueError, match="unsafe mapping selector"):
        load_mapping(machine_erp_ref="M1", parser_version="../generic-v1")


def test_explicit_unit_conversions_are_applied_once():
    mapping = build_column_map(["Cycle_Time_ms", "Inj_Pres_psi"])

    row = map_row({"Cycle_Time_ms": "2500", "Inj_Pres_psi": "2617"}, mapping)

    assert row["cycle_time_s"] == pytest.approx(2.5)
    assert row["peak_pressure_bar"] == pytest.approx(180.43, rel=0.01)
    assert mapping["Cycle_Time_ms"]["source_unit"] == "ms"
    assert mapping["Cycle_Time_ms"]["conversion_applied"] is True


def test_generic_canonical_dictionary_covers_process_fields():
    mapping = build_column_map([
        "switchover_position_mm", "cooling_time_s", "barrel_temp_zone2_c",
        "barrel_temp_zone3_c", "mold_temperature_c", "energy_kwh", "mold_open_time_s",
    ])

    assert {details["canonical"] for details in mapping.values()} == {
        "switchover_position", "cooling_time_s", "barrel_temp_zone2_c",
        "barrel_temp_zone3_c", "mold_temperature_c", "energy_kwh", "mold_open_time_s",
    }
