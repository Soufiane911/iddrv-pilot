"""Bounded, read-only structural inspection for industrial tabular exports.

The module deliberately separates a structural fingerprint from the byte hash
used for import idempotence. Fingerprints contain no path, file name, row
value, timestamp, or customer identifier from the data rows.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import unicodedata
import zipfile
from collections import defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable, Literal

import chardet


FINGERPRINT_VERSION = "iddrv.tabular-fingerprint/v1"
UNIT_TOKENS = {
    "s", "ms", "min", "h", "h:min", "bar", "pa", "mpa", "psi", "mm",
    "cm3", "cm³", "kn", "n", "c", "°c", "kwh", "%", "count", "-",
}
TIME_LABELS = {
    "time", "timestamp", "date", "heure", "datetime", "dateheure", "t007",
    "startedat", "debut", "début",
}
PROCESS_LABELS = {
    "cycletime", "cycletimes", "tcycle", "tempscycle", "t4012", "dosingtime", "tdos", "t4015",
    "injectiontime", "tinj", "t4018", "cushion", "cushionvolume", "vmat", "v4062",
    "switchoverpressure", "psw", "p4072", "peakpressure", "pmax", "cyclecounter", "ncycle", "f1403",
}
MACHINE_LABELS = {"machine", "machineid", "machineerpref", "refmachine", "réfmachine"}


class StructuralError(Exception):
    """Safe structural error whose text never contains source data or paths."""

    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code
        self.safe_message = message


@dataclass(frozen=True)
class ReaderLimits:
    max_file_bytes: int = 256 * 1024 * 1024
    max_sample_bytes: int = 1024 * 1024
    max_sample_rows: int = 2048
    max_columns: int = 2048
    max_cell_chars: int = 16384
    max_sheets: int = 32
    max_archive_entries: int = 10000
    max_archive_member_bytes: int = 256 * 1024 * 1024
    max_archive_uncompressed_bytes: int = 512 * 1024 * 1024
    max_compression_ratio: float = 250.0


@dataclass(frozen=True)
class StructuralInspection:
    source_format: str
    family: Literal["delimited_text", "spreadsheet", "binary"]
    encoding: str | None
    delimiter: str | None
    orientation: Literal["rows", "transposed", "unknown"]
    header_row: int | None
    unit_row: int | None
    data_start_row: int | None
    columns: tuple[str, ...]
    units: tuple[str | None, ...]
    metadata_keys: tuple[str, ...]
    reference_date: str | None
    sample_rows: tuple[tuple[Any, ...], ...]
    rows_observed: int
    truncated: bool
    sheet_name: str | None
    sheet_index: int | None
    sheet_count: int | None
    brand_hint: str
    brand_confidence: float


@dataclass(frozen=True)
class StructuralFingerprint:
    version: str
    algorithm: Literal["sha256"]
    digest: str
    payload: dict[str, Any]


@dataclass(frozen=True)
class StructuralClassification:
    label: Literal["machine_export", "tabular_unknown", "binary_unsupported"]
    score: float
    threshold: float
    signals: dict[str, bool]


def _normalize_label(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").replace("\ufeff", "")).casefold()
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[\s_.\-/\\]+", "", text.strip())


def _safe_cell(value: Any, limits: ReaderLimits) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    text = str(value)
    if len(text) > limits.max_cell_chars:
        raise StructuralError("INGEST_CELL_TOO_LARGE", "a cell exceeds the inspection limit")
    return value


def _check_source(path: Path, limits: ReaderLimits) -> int:
    if not path.exists() or not path.is_file():
        raise StructuralError("INGEST_FILE_NOT_FOUND", "source file is unavailable")
    size = path.stat().st_size
    if size == 0:
        raise StructuralError("INGEST_EMPTY_FILE", "source file is empty")
    if size > limits.max_file_bytes:
        raise StructuralError("INGEST_FILE_TOO_LARGE", "source file exceeds the inspection limit")
    return size


def _decode_sample(raw: bytes) -> tuple[str, str]:
    if raw.startswith((b"\xff\xfe", b"\xfe\xff")):
        encoding = "utf-16"
    elif raw.startswith(b"\xef\xbb\xbf"):
        encoding = "utf-8-sig"
    else:
        result = chardet.detect(raw[:8192])
        encoding = (result.get("encoding") or "utf-8").lower()
        encoding = {
            "ascii": "utf-8",
            "iso-8859-1": "latin-1",
            "windows-1252": "latin-1",
            "utf-16-le": "utf-16",
            "utf-16-be": "utf-16",
        }.get(encoding, encoding)
    try:
        return raw.decode(encoding, errors="strict"), encoding
    except (LookupError, UnicodeDecodeError) as exc:
        raise StructuralError("INGEST_TEXT_DECODE_FAILED", "source text cannot be decoded") from exc


def _utf16_without_bom(raw: bytes) -> str | None:
    sample = raw[:8192]
    if len(sample) < 8 or sample.startswith((b"\xff\xfe", b"\xfe\xff")):
        return None
    even = sample[0::2]
    odd = sample[1::2]
    even_nul = even.count(0) / max(1, len(even))
    odd_nul = odd.count(0) / max(1, len(odd))
    candidate = None
    if odd_nul > 0.3 and even_nul < 0.05:
        candidate = "utf-16-le"
    elif even_nul > 0.3 and odd_nul < 0.05:
        candidate = "utf-16-be"
    if candidate is None:
        return None
    try:
        decoded = sample.decode(candidate, errors="strict")
    except UnicodeDecodeError:
        return None
    printable = sum(char.isprintable() or char in "\r\n\t" for char in decoded) / max(1, len(decoded))
    has_table_shape = "\n" in decoded and any(delimiter in decoded for delimiter in (",", ";", "\t", "|"))
    return candidate if printable > 0.9 and has_table_shape else None


def _looks_binary(raw: bytes) -> bool:
    if raw.startswith((b"\xff\xfe", b"\xfe\xff", b"\xef\xbb\xbf")) or _utf16_without_bom(raw):
        return False
    sample = raw[:8192]
    if not sample:
        return False
    nul_ratio = sample.count(b"\x00") / len(sample)
    control_ratio = sum(byte < 9 or 13 < byte < 32 for byte in sample) / len(sample)
    return nul_ratio > 0.01 or control_ratio > 0.05


def _detect_delimiter(text: str) -> str:
    nonempty = [line for line in text.splitlines()[:50] if line.strip()]
    sample = "\n".join(nonempty)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        scores = {
            delimiter: (
                sum(line.count(delimiter) for line in nonempty[:20]),
                sum(line.count(delimiter) >= 2 for line in nonempty[:20]),
            )
            for delimiter in (";", "\t", ",", "|")
        }
        delimiter, score = max(scores.items(), key=lambda item: (item[1][1], item[1][0]))
        if score == (0, 0):
            # A one-column CSV is still a valid table; the delimiter is
            # structurally irrelevant but a stable default keeps the reader usable.
            return ","
        return delimiter


def _read_csv_matrix(text: str, delimiter: str, limits: ReaderLimits) -> tuple[list[list[Any]], bool]:
    matrix: list[list[Any]] = []
    reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)
    truncated = False
    try:
        for row_index, row in enumerate(reader):
            if row_index >= limits.max_sample_rows + 50:
                truncated = True
                break
            if len(row) > limits.max_columns:
                raise StructuralError("INGEST_TOO_MANY_COLUMNS", "table exceeds the column limit")
            matrix.append([_safe_cell(value, limits) for value in row])
    except csv.Error as exc:
        raise StructuralError("INGEST_TABLE_MALFORMED", "delimited text is malformed") from exc
    return matrix, truncated


def _header_score(row: Iterable[Any]) -> tuple[int, int]:
    row_values = list(row)
    cells = [str(value).strip() for value in row_values if value is not None and str(value).strip()]
    alpha = sum(bool(re.search(r"[A-Za-zÀ-ÿ]", value)) for value in cells)
    minimum_alpha = 1 if len(cells) == 1 else max(2, int(len(cells) * 0.35))
    if not cells or alpha < minimum_alpha:
        return (0, 0)
    return (len(row_values), alpha)


def _find_header(matrix: list[list[Any]]) -> int:
    candidates = [(index, _header_score(row)) for index, row in enumerate(matrix[:50])]
    candidates = [item for item in candidates if item[1] != (0, 0)]
    if not candidates:
        raise StructuralError("INGEST_HEADER_NOT_FOUND", "no tabular header was detected")
    return max(candidates, key=lambda item: (item[1][0], item[1][1], -item[0]))[0]


def _is_unit_row(row: Iterable[Any]) -> bool:
    cells = [_normalize_label(value) for value in row if value is not None and str(value).strip()]
    if not cells:
        return False
    known_units = {_normalize_label(token) for token in UNIT_TOKENS} - {""}
    units = sum(value in known_units for value in cells)
    return units >= 1 and units >= len(cells) * 0.25


def _detect_transposed(matrix: list[list[Any]], delimiter: str | None) -> bool:
    nonempty = [row for row in matrix if any(value is not None and str(value).strip() for value in row)]
    if len(nonempty) < 3:
        return False
    first = _normalize_label(nonempty[0][0] if nonempty[0] else "")
    if first in {"variable", "parameter", "parametre"}:
        return True
    first_column = {_normalize_label(row[0]) for row in nonempty[:50] if row}
    structural_labels = {
        "date", "heure", "time", "timestamp", "numcycle", "cyclecounter",
        "cycletime", "dosingtime", "injectiontime", "cushionvolume",
        "switchoverpressure", "peakpressure", "clampforce", "oiltemp",
    }
    widths = [len(row) for row in nonempty[:50]]
    average = sum(widths) / len(widths)
    return len(first_column & structural_labels) >= 2 and average > len(nonempty) * 2 and average > 30


def _metadata_summary(matrix: list[list[Any]], header_index: int) -> tuple[tuple[str, ...], str | None]:
    keys: list[str] = []
    reference_date = None
    for row in matrix[:header_index]:
        cells = [str(value).strip() for value in row if value is not None and str(value).strip()]
        if not cells:
            continue
        key = cells[0].split(":", 1)[0]
        normalized = _normalize_label(key)
        if normalized and not normalized.isdigit() and normalized not in keys:
            keys.append(normalized)
        if reference_date is None and normalized.startswith("date") and len(cells) > 1:
            reference_date = cells[1].split()[0]
    return tuple(sorted(keys)), reference_date


def _brand_hint(columns: Iterable[str], metadata_keys: Iterable[str]) -> tuple[str, float]:
    labels = {_normalize_label(label) for label in (*columns, *metadata_keys)}
    arburg = sum(label in {"t007", "t4012", "t4015", "t4018", "v4062", "p4072", "f1403"} for label in labels)
    if arburg >= 2:
        return "arburg", min(1.0, arburg / 5)
    if any("kistler" in label for label in labels):
        return "kistler", 0.75
    if any(label in {"tcycle", "activeplastificationtime", "cc300"} for label in labels):
        return "engel", 0.5
    return "generic", 0.0


def _inspection_from_matrix(
    matrix: list[list[Any]], *, source_format: str, family: Literal["delimited_text", "spreadsheet"],
    encoding: str | None, delimiter: str | None, truncated: bool, limits: ReaderLimits,
    sheet_name: str | None = None, sheet_index: int | None = None, sheet_count: int | None = None,
) -> StructuralInspection:
    if not matrix or not any(any(value is not None and str(value).strip() for value in row) for row in matrix):
        raise StructuralError("INGEST_EMPTY_TABLE", "source table is empty")
    transposed = _detect_transposed(matrix, delimiter)
    if transposed:
        parameter_rows = [row for row in matrix if row and str(row[0] or "").strip()]
        if parameter_rows and _normalize_label(parameter_rows[0][0]) in {"variable", "parameter", "parametre"}:
            parameter_rows = parameter_rows[1:]
        columns = tuple(str(row[0]).strip() for row in parameter_rows)
        width = max((len(row) for row in parameter_rows), default=1)
        cycle_limit = min(width - 1, limits.max_sample_rows)
        sample_rows = tuple(
            tuple(row[index] if index < len(row) else None for row in parameter_rows)
            for index in range(1, cycle_limit + 1)
        )
        brand, confidence = _brand_hint(columns, ())
        return StructuralInspection(
            source_format=source_format, family=family, encoding=encoding, delimiter=delimiter,
            orientation="transposed", header_row=0, unit_row=None, data_start_row=1,
            columns=columns, units=tuple(None for _ in columns), metadata_keys=(), reference_date=None,
            sample_rows=sample_rows, rows_observed=len(sample_rows),
            truncated=truncated or width - 1 > cycle_limit, sheet_name=sheet_name,
            sheet_index=sheet_index, sheet_count=sheet_count, brand_hint=brand,
            brand_confidence=round(confidence, 2),
        )

    header_index = _find_header(matrix)
    header_values = list(matrix[header_index])
    while header_values and (header_values[-1] is None or not str(header_values[-1]).strip()):
        header_values.pop()
    headers = tuple(
        str(value).strip() if value is not None and str(value).strip() else f"__unnamed_{index}"
        for index, value in enumerate(header_values)
    )
    if len(headers) > limits.max_columns:
        raise StructuralError("INGEST_TOO_MANY_COLUMNS", "table exceeds the column limit")
    unit_index = header_index + 1 if header_index + 1 < len(matrix) and _is_unit_row(matrix[header_index + 1]) else None
    data_start = (unit_index + 1) if unit_index is not None else header_index + 1
    units_row = matrix[unit_index] if unit_index is not None else []
    units = tuple(
        str(units_row[index]).strip() if index < len(units_row) and units_row[index] is not None else None
        for index in range(len(headers))
    )
    data = []
    for row in matrix[data_start:data_start + limits.max_sample_rows]:
        normalized = tuple(row[index] if index < len(row) else None for index in range(len(headers)))
        if any(value is not None and str(value).strip() for value in normalized):
            data.append(normalized)
    metadata, reference_date = _metadata_summary(matrix, header_index)
    brand, confidence = _brand_hint(headers, metadata)
    return StructuralInspection(
        source_format=source_format, family=family, encoding=encoding, delimiter=delimiter,
        orientation="rows", header_row=header_index, unit_row=unit_index,
        data_start_row=data_start, columns=headers, units=units, metadata_keys=metadata,
        reference_date=reference_date, sample_rows=tuple(data), rows_observed=len(data),
        truncated=truncated or len(matrix) > data_start + limits.max_sample_rows,
        sheet_name=sheet_name, sheet_index=sheet_index, sheet_count=sheet_count,
        brand_hint=brand, brand_confidence=round(confidence, 2),
    )


def _inspect_text(path: Path, source_format: str, limits: ReaderLimits) -> StructuralInspection:
    with path.open("rb") as handle:
        raw = handle.read(limits.max_sample_bytes + 1)
    sample_truncated = len(raw) > limits.max_sample_bytes
    raw = raw[:limits.max_sample_bytes]
    utf16_encoding = _utf16_without_bom(raw)
    if _looks_binary(raw):
        if source_format == "dat":
            return StructuralInspection(
                source_format="dat", family="binary", encoding=None, delimiter=None,
                orientation="unknown", header_row=None, unit_row=None, data_start_row=None,
                columns=(), units=(), metadata_keys=(), reference_date=None, sample_rows=(), rows_observed=0,
                truncated=sample_truncated, sheet_name=None, sheet_index=None, sheet_count=None,
                brand_hint="generic", brand_confidence=0.0,
            )
        raise StructuralError("INGEST_BINARY_TEXT", "binary content is not a tabular text export")
    if utf16_encoding:
        try:
            text, encoding = raw.decode(utf16_encoding, errors="strict"), utf16_encoding
        except UnicodeDecodeError as exc:
            raise StructuralError("INGEST_TEXT_DECODE_FAILED", "source text cannot be decoded") from exc
    else:
        text, encoding = _decode_sample(raw)
    delimiter = _detect_delimiter(text)
    matrix, row_truncated = _read_csv_matrix(text, delimiter, limits)
    return _inspection_from_matrix(
        matrix, source_format=source_format, family="delimited_text", encoding=encoding,
        delimiter=delimiter, truncated=sample_truncated or row_truncated, limits=limits,
    )


def _preflight_xlsx(path: Path, limits: ReaderLimits) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            infos = archive.infolist()
            if len(infos) > limits.max_archive_entries:
                raise StructuralError("INGEST_ARCHIVE_ENTRIES", "spreadsheet archive has too many entries")
            for info in infos:
                if info.file_size > limits.max_archive_member_bytes:
                    raise StructuralError("INGEST_ARCHIVE_MEMBER_TOO_LARGE", "spreadsheet archive member exceeds the limit")
                if info.file_size / max(1, info.compress_size) > limits.max_compression_ratio:
                    raise StructuralError("INGEST_ARCHIVE_RATIO", "spreadsheet compression ratio is unsafe")
            total_uncompressed = sum(info.file_size for info in infos)
            total_compressed = sum(max(1, info.compress_size) for info in infos)
            if total_uncompressed > limits.max_archive_uncompressed_bytes:
                raise StructuralError("INGEST_ARCHIVE_TOO_LARGE", "spreadsheet archive exceeds the inspection limit")
            if total_uncompressed / max(1, total_compressed) > limits.max_compression_ratio:
                raise StructuralError("INGEST_ARCHIVE_RATIO", "spreadsheet compression ratio is unsafe")
    except zipfile.BadZipFile as exc:
        raise StructuralError("INGEST_SPREADSHEET_READ_FAILED", "spreadsheet container is invalid") from exc


def _worksheet_matrix(ws: Any, limits: ReaderLimits) -> tuple[list[list[Any]], bool]:
    if ws.max_column is not None and ws.max_column > limits.max_columns:
        raise StructuralError("INGEST_TOO_MANY_COLUMNS", "table exceeds the column limit")
    matrix: list[list[Any]] = []
    truncated = bool(ws.max_row is not None and ws.max_row > limits.max_sample_rows + 50)
    for row_index, row in enumerate(ws.iter_rows(
        min_row=1,
        max_row=limits.max_sample_rows + 50,
        max_col=limits.max_columns + 1,
        values_only=True,
    )):
        if row_index >= limits.max_sample_rows + 50:
            truncated = True
            break
        values = [_safe_cell(value, limits) for value in row]
        while values and values[-1] is None:
            values.pop()
        if len(values) > limits.max_columns:
            raise StructuralError("INGEST_TOO_MANY_COLUMNS", "table exceeds the column limit")
        matrix.append(values)
    return matrix, truncated


def _inspect_xlsx(path: Path, limits: ReaderLimits, sheet_name: str | None) -> StructuralInspection:
    _preflight_xlsx(path, limits)
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        try:
            if len(workbook.sheetnames) > limits.max_sheets:
                raise StructuralError("INGEST_TOO_MANY_SHEETS", "spreadsheet exceeds the sheet limit")
            if sheet_name is not None:
                if sheet_name not in workbook.sheetnames:
                    raise StructuralError("INGEST_SHEET_NOT_FOUND", "requested spreadsheet sheet is unavailable")
                selected_index = workbook.sheetnames.index(sheet_name)
            else:
                scored: list[tuple[tuple[int, int], int]] = []
                for index, name in enumerate(workbook.sheetnames):
                    worksheet = workbook[name]
                    if worksheet.max_column is not None and worksheet.max_column > limits.max_columns:
                        raise StructuralError("INGEST_TOO_MANY_COLUMNS", "table exceeds the column limit")
                    preview = []
                    for row in worksheet.iter_rows(
                        min_row=1, max_row=50, max_col=limits.max_columns + 1, values_only=True,
                    ):
                        values = [value for value in row if value is not None and str(value).strip()]
                        preview.append(values)
                    score = max((_header_score(row) for row in preview), default=(0, 0))
                    scored.append((score, index))
                selected_index = max(scored, key=lambda item: (item[0], -item[1]))[1]
            selected_name = workbook.sheetnames[selected_index]
            matrix, truncated = _worksheet_matrix(workbook[selected_name], limits)
            return _inspection_from_matrix(
                matrix, source_format="xlsx", family="spreadsheet", encoding=None,
                delimiter=None, truncated=truncated, limits=limits, sheet_name=selected_name,
                sheet_index=selected_index, sheet_count=len(workbook.sheetnames),
            )
        finally:
            workbook.close()
    except StructuralError:
        raise
    except Exception as exc:
        raise StructuralError("INGEST_SPREADSHEET_READ_FAILED", "spreadsheet cannot be inspected") from exc


def _inspect_xls(path: Path, limits: ReaderLimits, sheet_name: str | None) -> StructuralInspection:
    try:
        import pandas as pd
        import xlrd  # noqa: F401 - explicit capability check
    except ImportError as exc:
        raise StructuralError("INGEST_XLS_ENGINE_UNAVAILABLE", "legacy spreadsheet engine is unavailable") from exc
    try:
        book = pd.ExcelFile(path, engine="xlrd")
        if len(book.sheet_names) > limits.max_sheets:
            raise StructuralError("INGEST_TOO_MANY_SHEETS", "spreadsheet exceeds the sheet limit")
        if sheet_name is not None:
            if sheet_name not in book.sheet_names:
                raise StructuralError("INGEST_SHEET_NOT_FOUND", "requested spreadsheet sheet is unavailable")
            selected_index = book.sheet_names.index(sheet_name)
        else:
            scores = []
            for index, name in enumerate(book.sheet_names):
                frame = pd.read_excel(book, sheet_name=name, header=None, nrows=50)
                scores.append((max((_header_score(row) for row in frame.itertuples(index=False, name=None)), default=(0, 0)), index))
            selected_index = max(scores, key=lambda item: (item[0], -item[1]))[1]
        selected_name = book.sheet_names[selected_index]
        frame = pd.read_excel(book, sheet_name=selected_name, header=None, nrows=limits.max_sample_rows + 50)
        matrix = [[_safe_cell(value, limits) for value in row] for row in frame.itertuples(index=False, name=None)]
        return _inspection_from_matrix(
            matrix, source_format="xls", family="spreadsheet", encoding=None, delimiter=None,
            truncated=len(frame) >= limits.max_sample_rows + 50, limits=limits,
            sheet_name=selected_name, sheet_index=selected_index, sheet_count=len(book.sheet_names),
        )
    except StructuralError:
        raise
    except Exception as exc:
        raise StructuralError("INGEST_SPREADSHEET_READ_FAILED", "legacy spreadsheet cannot be inspected") from exc


def inspect_file(
    path: str | Path, *, limits: ReaderLimits = ReaderLimits(), sheet_name: str | None = None,
) -> StructuralInspection:
    source = Path(path)
    _check_source(source, limits)
    extension = source.suffix.lower()
    if extension in {".csv", ".txt", ".dat"}:
        return _inspect_text(source, extension[1:], limits)
    if extension == ".xlsx":
        return _inspect_xlsx(source, limits, sheet_name)
    if extension == ".xls":
        return _inspect_xls(source, limits, sheet_name)
    raise StructuralError("INGEST_UNSUPPORTED_FORMAT", "source format is unsupported")


def fingerprint_inspection(inspection: StructuralInspection) -> StructuralFingerprint:
    if inspection.family == "binary":
        payload: dict[str, Any] = {
            "schema": FINGERPRINT_VERSION,
            "family": "binary",
            "source_format": inspection.source_format,
            "classification": "unsupported",
        }
    else:
        payload = {
            "schema": FINGERPRINT_VERSION,
            "family": inspection.family,
            "orientation": inspection.orientation,
            "delimiter": inspection.delimiter,
            "header_row": inspection.header_row,
            "unit_row": inspection.unit_row,
            "data_start_row": inspection.data_start_row,
            "columns": [
                {
                    "position": index,
                    "label": _normalize_label(label),
                    "unit": _normalize_label(inspection.units[index]) if index < len(inspection.units) and inspection.units[index] else None,
                }
                for index, label in enumerate(inspection.columns)
            ],
            "metadata_keys": list(inspection.metadata_keys),
        }
    canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return StructuralFingerprint(
        version=FINGERPRINT_VERSION,
        algorithm="sha256",
        digest=hashlib.sha256(canonical.encode("utf-8")).hexdigest(),
        payload=payload,
    )


def fingerprint_file(
    path: str | Path, *, limits: ReaderLimits = ReaderLimits(), sheet_name: str | None = None,
) -> StructuralFingerprint:
    return fingerprint_inspection(inspect_file(path, limits=limits, sheet_name=sheet_name))


def classify_structure(inspection: StructuralInspection) -> StructuralClassification:
    if inspection.family == "binary":
        return StructuralClassification(
            label="binary_unsupported", score=0.0, threshold=0.8,
            signals={"timestamp": False, "process_metric": False, "machine_context": False},
        )
    labels = {_normalize_label(label) for label in inspection.columns}
    has_time = bool(labels & {_normalize_label(label) for label in TIME_LABELS}) or {"date", "heure"}.issubset(labels)
    has_process = bool(labels & {_normalize_label(label) for label in PROCESS_LABELS})
    has_context = bool(labels & {_normalize_label(label) for label in MACHINE_LABELS}) or inspection.brand_hint != "generic"
    score = round(0.4 * has_time + 0.4 * has_process + 0.2 * has_context, 2)
    return StructuralClassification(
        label="machine_export" if score >= 0.8 else "tabular_unknown",
        score=score, threshold=0.8,
        signals={"timestamp": has_time, "process_metric": has_process, "machine_context": has_context},
    )


def group_by_fingerprint(
    paths: Iterable[str | Path], *, limits: ReaderLimits = ReaderLimits(),
) -> dict[str, tuple[Path, ...]]:
    groups: dict[str, list[Path]] = defaultdict(list)
    for path in paths:
        source = Path(path)
        groups[fingerprint_file(source, limits=limits).digest].append(source)
    return {digest: tuple(items) for digest, items in sorted(groups.items())}


__all__ = [
    "FINGERPRINT_VERSION", "ReaderLimits", "StructuralClassification", "StructuralError",
    "StructuralFingerprint", "StructuralInspection", "classify_structure", "fingerprint_file",
    "fingerprint_inspection", "group_by_fingerprint", "inspect_file",
]
